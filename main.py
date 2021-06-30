__author__ = 'Eric Hom'
__copyright__ = 'Copyright (c) 2021. All rights are reserved.'
__license__ = 'GPL 3'

import fitz
from PyQt5 import QtWidgets
from openpyxl import Workbook
from pathlib import Path
from ui import Ui_Dialog
from PyQt5.QtWidgets import QDialog, QApplication


class Dialog(QDialog, Ui_Dialog):
    """
    Extracts all comments from PDFs to an excel sheet
    """
    def __init__(self, parent=None):
        super(Dialog, self).__init__(parent)
        self.setupUi(self)
        self.buttonBox.accepted.connect(self.buttonClicked)  # alternative way to call your method

    def buttonClicked(self):
        pdf_path = Path(self.pdfPath.text())
        counter = 1
        comment_list=[]
        for file in pdf_path.iterdir():
            if file.name.endswith(".pdf"):
                pdf = fitz.open(str(file))
                for i in range(pdf.pageCount):
                    page = pdf[i]
                    for annot in page.annots():
                        str_comment = str(annot.info["content"])
                        row_data= (counter,"Open", "Round 1",file.name,i+1,"","","","",str_comment,"Yes","","")
                        counter += 1
                        comment_list.append(row_data)

        row_count = 2
        wb = Workbook()
        ws = wb.active
        ws['A1'] = "Issue #"
        ws['B1'] = "Status"
        ws['C1'] = "Testing Round"
        ws['D1'] = "File Name"
        ws['E1'] = "PDF Page"
        ws['F1'] = "TWB Line"
        ws['G1'] = "Source Text"
        ws['H1'] = "Current Translation"
        ws['I1'] = "Updated Translation"
        ws['J1'] = "Issue Description"
        ws['K1'] = "Implemented in TWB"
        ws['L1'] = "MDSol Comment"
        ws['M1'] = "Vendor Feedback"
        for entry in comment_list:
            column = 1
            for i in range(0,12):
                ws.cell(row=row_count,column=column).value = entry[i]
                column+=1
            row_count +=1
        print(pdf_path.parent)
        wb.save((self.pdfPath.text()+"PDF_Comment_Report.xlsx"))


if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    macro_dialog = Dialog() # create object of dialog, **use the name of your class (ie class Dialog)**
    macro_dialog.show()
    sys.exit(app.exec_())

